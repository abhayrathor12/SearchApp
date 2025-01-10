import pandas as pd
from searchlistapp.models import Productlist  # Adjust import as per your app
from django.shortcuts import render
from django.http import JsonResponse, HttpResponse
import openpyxl
from openpyxl.utils import get_column_letter
import json
from django.views.decorators.csrf import csrf_exempt
import pandas as pd
from django.http import HttpResponse
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font


def homepage(request):

    return render(request, "home.html")


def ajax_search(request):
    query = request.GET.get("q", "")  # Get the query string
    results = []
    if query:
        products = Productlist.objects.filter(
            model__icontains=query
        )  # Filter models containing the query
        results = [
            {"model": product.model, "list_price": str(product.list_price)}
            for product in products
        ]
    return JsonResponse({"results": results})


@csrf_exempt
def download_excel(request):
    if request.method == "POST":
        # Parse the JSON data sent from the frontend
        request_data = json.loads(request.body)
        data_new = request_data.get("data", [])  # Get the product data
        include_discount = data_new.get("include_discount")
        include_list_price = data_new.get("include_list_price")
        data_list = data_new.get("data", [])
        print(data_new)
        # Print the received data for debugging
        print(f"Data received: {data_list}", include_discount, include_list_price)

        # Initialize a variable to calculate the subtotal sum
        total_subtotal = 0

        # Clean up the data by removing unnecessary columns
        for row in data_list:
            del row["Total"]
            if not include_discount and "Discount" in row:
                del row["Discount"]
            if not include_list_price and "Price" in row:
                del row["Price"]

            # Accumulate the subtotal (assuming there's a 'Subtotal' key in the row)
            if "SubTotal" in row:
                try:
                    total_subtotal += float(row["SubTotal"])  # Add to the total sum
                except ValueError:
                    pass  # If the value is not a valid number, just skip

        print(data_list, "yryyyr")

        # Create a new Excel workbook and sheet using openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Catalog Data"

        # Define styles
        header_fill = PatternFill(
            start_color="FFFF00", end_color="FFFF00", fill_type="solid"
        )
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )
        center_alignment = Alignment(horizontal="center", vertical="center")

        # Define all headers dynamically from the first dictionary in data_list
        headers = (
            list(data_list[0].keys()) if data_list else []
        )  # Extract keys from the first item
        if headers:
            for col_num, header in enumerate(headers, 1):
                cell = ws[get_column_letter(col_num) + "1"]
                cell.value = header
                cell.fill = header_fill  # Apply yellow background
                cell.border = thin_border  # Apply border
                cell.alignment = center_alignment  # Center-align text
                cell.font = Font(bold=True)  # Make header bold

        # Write the data rows
        for row_num, item in enumerate(data_list, 2):  # Start from row 2
            for col_num, header in enumerate(headers, 1):
                cell = ws[get_column_letter(col_num) + str(row_num)]
                cell.value = item.get(header, "")  # Write data
                cell.border = thin_border  # Apply border
                cell.alignment = center_alignment  # Center-align text

        # Add the Total Row at the bottom
        total_row_num = len(data_list) + 2  # Row number for the total
        ws[f"A{total_row_num}"].value = "Total"  # Label the first cell as "Total"
        ws[f"A{total_row_num}"].font = Font(bold=True)  # Make it bold
        ws[f"A{total_row_num}"].border = thin_border
        ws[f"A{total_row_num}"].alignment = center_alignment  # Center-align the text

        # Write the total for the 'Subtotal' column (assuming it is at the index corresponding to the 'Subtotal' header)
        subtotal_col_index = (
            headers.index("SubTotal") + 1
        )  # Index of the 'Subtotal' column (1-based)
        ws[get_column_letter(subtotal_col_index) + str(total_row_num)].value = (
            total_subtotal
        )
        ws[get_column_letter(subtotal_col_index) + str(total_row_num)].font = Font(
            bold=True
        )  # Make the total bold
        ws[get_column_letter(subtotal_col_index) + str(total_row_num)].border = (
            thin_border
        )
        ws[get_column_letter(subtotal_col_index) + str(total_row_num)].alignment = (
            center_alignment  # Center-align the total value
        )

        # Adjust column width based on the content
        for col_num, header in enumerate(headers, 1):
            column_letter = get_column_letter(col_num)
            ws.column_dimensions[column_letter].width = max(len(header), 15)

        # Apply outer border around the entire table
        max_row = len(data_list) + 2  # Including the header and total row
        max_col = len(headers)

        # Apply outer border for all data and the total row
        outer_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        for row in range(1, max_row + 1):
            for col in range(1, max_col + 1):
                cell = ws[get_column_letter(col) + str(row)]
                cell.border = outer_border  # Apply outer border to all cells

        # Prepare the response as an Excel file
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = 'attachment; filename="CatalogData.xlsx"'

        # Save the Excel file to the response
        wb.save(response)

        return response

    return JsonResponse({"error": "Invalid request"}, status=400)


@csrf_exempt
def download_excel_writer(request):
    if request.method == "POST":
        # Get data from the frontend (in JSON format)
        data = json.loads(request.body).get("data", [])
        print(type(data), data)
        # Create a new Excel workbook and sheet
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Catalog Data"

        # Write the header
        headers = ["Catalog", "Price", "Discount"]
        for col_num, header in enumerate(headers, 1):
            ws[get_column_letter(col_num) + "1"] = header

        # Write the data rows
        for row_num, item in enumerate(data, 2):
            ws["A" + str(row_num)] = item.get("Catalog")
            ws["B" + str(row_num)] = item.get("Price")
            ws["C" + str(row_num)] = item.get("Discount")

        # Prepare the response as an Excel file
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = "attachment; filename=data.xlsx"

        # Save the Excel file to the response
        wb.save(response)

        return response
    return JsonResponse({"error": "Invalid request"}, status=400)


# def search_products(request):

#     return render(request, "search_results.html", {"query": query, "results": results})


# def import_productlist_from_csv(csv_path):
#     # Load CSV data
#     data = pd.read_csv(csv_path, skiprows=1, names=["model", "list_price"])

#     # Clean data
#     data = data.dropna()  # Remove rows with missing values
#     data["list_price"] = (
#         data["list_price"]
#         .str.replace(r"[₹,]", "", regex=True)  # Remove currency symbols and commas
#         .str.extract(r"(\d+\.?\d*)")[0]  # Extract numeric values
#     )

#     # Ensure valid numeric conversion
#     data = data.dropna(subset=["list_price"])  # Drop rows where list_price is NaN
#     data["list_price"] = data["list_price"].astype(float)  # Convert to float

#     # Iterate through rows and save to database
#     for _, row in data.iterrows():
#         Productlist.objects.create(model=row["model"], list_price=row["list_price"])

#     print("Data imported successfully.")


# # File path
# csv_path = "C:/Users/DELL/Desktop/SearchApp/searchlist/output.csv"
# import_productlist_from_csv(csv_path)
